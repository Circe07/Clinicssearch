#!/usr/bin/env python3
"""
Buscador de clínicas sin web o con web básica.

Busca clínicas en una zona determinada usando la API de Google Places,
analiza si tienen página web y evalúa su calidad. Exporta los resultados
a un archivo Excel con: nombre, teléfono, email, dirección y enlace web.

Uso:
    python buscar_clinicas.py --query "clínicas dentales en Madrid"
    python buscar_clinicas.py --query "clínicas estéticas en Barcelona" --radius 10000
    python buscar_clinicas.py --query "fisioterapia en Valencia" --output mis_clinicas.xlsx
"""

import argparse
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PLACES_TEXT_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"

SEARCH_FIELD_MASK = ",".join([
    "places.id",
    "places.displayName",
    "places.formattedAddress",
    "places.nationalPhoneNumber",
    "places.internationalPhoneNumber",
    "places.websiteUri",
    "places.googleMapsUri",
    "places.rating",
    "places.userRatingCount",
    "nextPageToken",
])

HEADERS_BROWSER = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
}

WEB_QUALITY_THRESHOLDS = {
    "min_pages_links": 5,
    "min_content_length": 3000,
    "modern_indicators": [
        "react", "vue", "angular", "next", "nuxt", "gatsby",
        "tailwind", "bootstrap", "wordpress", "wp-content",
        "shopify", "squarespace", "wix",
    ],
    "basic_indicators": [
        "under construction", "en construcción", "coming soon",
        "próximamente", "página en desarrollo", "sitio en mantenimiento",
    ],
}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Busca clínicas sin web o con web básica y exporta a Excel."
    )
    parser.add_argument(
        "--query", "-q",
        required=True,
        help='Búsqueda de Google Places, ej: "clínicas dentales en Madrid"',
    )
    parser.add_argument(
        "--api-key", "-k",
        required=True,
        help="Tu API key de Google Places.",
    )
    parser.add_argument(
        "--radius", "-r",
        type=int,
        default=5000,
        help="Radio de búsqueda en metros (default: 5000).",
    )
    parser.add_argument(
        "--max-results", "-m",
        type=int,
        default=60,
        help="Número máximo de resultados (default: 60, max ~60 con paginación).",
    )
    parser.add_argument(
        "--output", "-o",
        default="clinicas_prospecto.xlsx",
        help="Nombre del archivo Excel de salida (default: clinicas_prospecto.xlsx).",
    )
    parser.add_argument(
        "--workers", "-w",
        type=int,
        default=5,
        help="Hilos paralelos para analizar webs (default: 5).",
    )
    parser.add_argument(
        "--min-reviews", "-n",
        type=int,
        default=50,
        help="Mínimo de reseñas para incluir una clínica (default: 50).",
    )
    return parser.parse_args()


def search_places(query: str, api_key: str, max_results: int = 60) -> list[dict]:
    """Busca lugares con Google Places API (New) Text Search (con paginación)."""
    results = []
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": SEARCH_FIELD_MASK,
    }
    page_size = min(max_results, 20)
    body = {
        "textQuery": query,
        "languageCode": "es",
        "pageSize": page_size,
    }

    while len(results) < max_results:
        resp = requests.post(PLACES_TEXT_SEARCH_URL, headers=headers, json=body, timeout=15)

        if resp.status_code != 200:
            error_data = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
            error_msg = error_data.get("error", {}).get("message", resp.text[:200])
            error_status = error_data.get("error", {}).get("status", resp.status_code)
            print(f"[ERROR] Google Places API: {error_status} - {error_msg}")
            if resp.status_code == 403 or "API_KEY" in str(error_msg).upper():
                print("        Verifica que tu API key sea válida y tenga habilitada la 'Places API (New)'.")
                print("        En Google Cloud Console > APIs y servicios > habilita 'Places API (New)'.")
            break

        data = resp.json()
        places = data.get("places", [])
        if not places:
            break

        results.extend(places)
        next_token = data.get("nextPageToken")
        if not next_token or len(results) >= max_results:
            break

        body = {
            "textQuery": query,
            "languageCode": "es",
            "pageSize": page_size,
            "pageToken": next_token,
        }

    return results[:max_results]


def extract_emails_from_html(html: str) -> list[str]:
    """Extrae emails del HTML de una página."""
    pattern = r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
    raw = re.findall(pattern, html)
    cleaned = []
    seen = set()
    for email in raw:
        email_lower = email.lower()
        if email_lower not in seen and not email_lower.endswith((".png", ".jpg", ".gif", ".svg", ".webp")):
            seen.add(email_lower)
            cleaned.append(email)
    return cleaned


def analyze_website(url: str) -> dict:
    """
    Analiza la calidad de un sitio web.
    Retorna un dict con: accesible, calidad, emails, descripción.
    """
    result = {
        "accesible": False,
        "calidad": "Sin web",
        "emails": [],
        "descripcion": "No tiene página web",
        "num_links": 0,
        "contenido_length": 0,
    }

    if not url:
        return result

    try:
        resp = requests.get(
            url,
            headers=HEADERS_BROWSER,
            timeout=10,
            allow_redirects=True,
            verify=True,
        )
        resp.raise_for_status()
    except requests.exceptions.SSLError:
        try:
            resp = requests.get(url, headers=HEADERS_BROWSER, timeout=10, allow_redirects=True, verify=False)
            resp.raise_for_status()
        except Exception:
            result["calidad"] = "Web inaccesible"
            result["descripcion"] = "El sitio no responde o tiene errores SSL"
            return result
    except requests.exceptions.ConnectionError:
        result["calidad"] = "Web inaccesible"
        result["descripcion"] = "No se pudo conectar al servidor"
        return result
    except requests.exceptions.Timeout:
        result["calidad"] = "Web inaccesible"
        result["descripcion"] = "Tiempo de espera agotado"
        return result
    except requests.exceptions.RequestException:
        result["calidad"] = "Web inaccesible"
        result["descripcion"] = "Error al acceder a la web"
        return result

    result["accesible"] = True
    html = resp.text
    soup = BeautifulSoup(html, "html.parser")

    text_content = soup.get_text(separator=" ", strip=True)
    result["contenido_length"] = len(text_content)

    internal_links = set()
    parsed_base = urlparse(url)
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        parsed_href = urlparse(href)
        if parsed_href.netloc == "" or parsed_href.netloc == parsed_base.netloc:
            internal_links.add(href)
    result["num_links"] = len(internal_links)

    result["emails"] = extract_emails_from_html(html)

    html_lower = html.lower()

    for indicator in WEB_QUALITY_THRESHOLDS["basic_indicators"]:
        if indicator in html_lower or indicator in text_content.lower():
            result["calidad"] = "MUY BÁSICA / En construcción"
            result["descripcion"] = f"Sitio en construcción o placeholder ('{indicator}')"
            return result

    has_modern_tech = any(ind in html_lower for ind in WEB_QUALITY_THRESHOLDS["modern_indicators"])
    few_links = result["num_links"] < WEB_QUALITY_THRESHOLDS["min_pages_links"]
    short_content = result["contenido_length"] < WEB_QUALITY_THRESHOLDS["min_content_length"]

    if short_content and few_links:
        result["calidad"] = "MUY BÁSICA"
        result["descripcion"] = (
            f"Contenido escaso ({result['contenido_length']} chars) "
            f"y pocas secciones ({result['num_links']} links internos)"
        )
    elif short_content or few_links:
        result["calidad"] = "BÁSICA"
        result["descripcion"] = (
            f"Contenido: {result['contenido_length']} chars, "
            f"Links internos: {result['num_links']}"
        )
    elif has_modern_tech:
        result["calidad"] = "PROFESIONAL"
        result["descripcion"] = "Sitio con tecnología moderna y contenido adecuado"
    else:
        result["calidad"] = "ACEPTABLE"
        result["descripcion"] = (
            f"Contenido: {result['contenido_length']} chars, "
            f"Links internos: {result['num_links']} (sin tecnología moderna detectada)"
        )

    return result


def process_clinic(place: dict) -> dict:
    """Procesa una clínica: extrae datos y analiza su web."""
    display_name = place.get("displayName", {})
    name = display_name.get("text", "Sin nombre") if isinstance(display_name, dict) else str(display_name)
    address = place.get("formattedAddress", "")
    phone = place.get("nationalPhoneNumber", "") or place.get("internationalPhoneNumber", "")
    website = place.get("websiteUri", "")
    google_maps_url = place.get("googleMapsUri", "")
    rating = place.get("rating", "")
    total_ratings = place.get("userRatingCount", "")

    web_analysis = analyze_website(website)

    emails = web_analysis["emails"]
    email_str = ", ".join(emails) if emails else ""

    return {
        "nombre": name,
        "telefono": phone,
        "email": email_str,
        "direccion": address,
        "web": website,
        "google_maps": google_maps_url,
        "calidad_web": web_analysis["calidad"],
        "descripcion_web": web_analysis["descripcion"],
        "rating": rating,
        "num_reviews": total_ratings,
        "oportunidad": web_analysis["calidad"] in ("Sin web", "Web inaccesible", "MUY BÁSICA", "MUY BÁSICA / En construcción", "BÁSICA"),
    }


def _clinic_key(nombre: str, direccion: str) -> str:
    """Genera una clave única para identificar duplicados (nombre + dirección normalizados)."""
    return (nombre.strip().lower() + "|" + direccion.strip().lower())


def load_existing_clinics(output_path: str) -> list[dict]:
    """Carga clínicas ya guardadas en un Excel existente."""
    if not os.path.exists(output_path):
        return []

    try:
        wb = load_workbook(output_path, read_only=True)
    except Exception as e:
        print(f"[AVISO] No se pudo leer el Excel existente ({e}). Se creará uno nuevo.")
        return []

    if "Clínicas - Prospectos" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Clínicas - Prospectos"]
    clinics = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue
        clinics.append({
            "nombre": row[0] or "",
            "telefono": row[1] or "",
            "email": row[2] or "",
            "direccion": row[3] or "",
            "web": row[4] or "",
            "google_maps": row[5] or "",
            "calidad_web": row[6] or "",
            "descripcion_web": row[7] or "",
            "rating": row[8] if len(row) > 8 and row[8] is not None else "",
            "num_reviews": row[9] if len(row) > 9 and row[9] is not None else "",
            "oportunidad": True,
        })

    wb.close()
    return clinics


def merge_clinics(existing: list[dict], new: list[dict]) -> tuple[list[dict], int, int]:
    """
    Fusiona clínicas existentes con nuevas, sin duplicar.
    Retorna (lista_final, num_nuevas_añadidas, num_duplicadas_ignoradas).
    """
    seen_keys = set()
    for clinic in existing:
        seen_keys.add(_clinic_key(clinic["nombre"], clinic["direccion"]))

    merged = list(existing)
    added = 0
    skipped = 0

    for clinic in new:
        key = _clinic_key(clinic["nombre"], clinic["direccion"])
        if key in seen_keys:
            skipped += 1
        else:
            seen_keys.add(key)
            merged.append(clinic)
            added += 1

    return merged, added, skipped


def save_to_excel(clinics: list[dict], output_path: str):
    """Guarda los resultados en un archivo Excel con formato profesional."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clínicas - Prospectos"

    headers = [
        "Nombre",
        "Teléfono",
        "Email",
        "Dirección",
        "Página Web",
        "Google Maps",
        "Calidad Web",
        "Descripción",
        "Rating",
        "Nº Reseñas",
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    oportunidad_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    sorted_clinics = sorted(clinics, key=lambda c: (c["calidad_web"], c["nombre"]))

    for row_idx, clinic in enumerate(sorted_clinics, 2):
        values = [
            clinic["nombre"],
            clinic["telefono"],
            clinic["email"],
            clinic["direccion"],
            clinic["web"],
            clinic["google_maps"],
            clinic["calidad_web"],
            clinic["descripcion_web"],
            clinic["rating"],
            clinic["num_reviews"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col_idx in (5, 6) and value:
                cell.font = Font(color="0563C1", underline="single")

    col_widths = {1: 35, 2: 18, 3: 30, 4: 45, 5: 40, 6: 40, 7: 22, 8: 50, 9: 10, 10: 12}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    if "Resumen" in wb.sheetnames:
        del wb["Resumen"]
    ws_stats = wb.create_sheet("Resumen")
    stats_header_font = Font(name="Calibri", size=12, bold=True)
    stats_data = [
        ("Resumen de Oportunidades", ""),
        ("", ""),
        ("Total oportunidades guardadas", len(clinics)),
        ("", ""),
        ("Sin página web", sum(1 for c in clinics if c["calidad_web"] == "Sin web")),
        ("Web inaccesible", sum(1 for c in clinics if c["calidad_web"] == "Web inaccesible")),
        ("Web muy básica / en construcción", sum(1 for c in clinics if "MUY BÁSICA" in c["calidad_web"])),
        ("Web básica", sum(1 for c in clinics if c["calidad_web"] == "BÁSICA")),
    ]

    for row_idx, (label, value) in enumerate(stats_data, 1):
        cell_a = ws_stats.cell(row=row_idx, column=1, value=label)
        cell_b = ws_stats.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = stats_header_font
        if row_idx == len(stats_data):
            cell_a.font = Font(bold=True, size=12, color="006100")
            cell_b.font = Font(bold=True, size=12, color="006100")
            cell_a.fill = oportunidad_fill
            cell_b.fill = oportunidad_fill

    ws_stats.column_dimensions["A"].width = 35
    ws_stats.column_dimensions["B"].width = 15

    wb.save(output_path)


def main():
    args = parse_args()

    print("=" * 60)
    print("  BUSCADOR DE CLÍNICAS - Detector de Oportunidades Web")
    print("=" * 60)
    print(f"\n  Búsqueda: {args.query}")
    print(f"  Max resultados: {args.max_results}")
    print(f"  Mín. reseñas: {args.min_reviews}")
    print(f"  Archivo de salida: {args.output}")
    print()

    print("[1/4] Buscando clínicas en Google Places...")
    places = search_places(args.query, args.api_key, args.max_results)

    if not places:
        print("\n  No se encontraron resultados. Intenta con otra búsqueda.")
        sys.exit(1)

    print(f"      Se encontraron {len(places)} clínicas.")

    print(f"\n[2/4] Analizando webs de las clínicas ({args.workers} hilos)...")
    clinics_raw = []
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(process_clinic, place): place
            for place in places
        }
        for i, future in enumerate(as_completed(futures), 1):
            try:
                clinic = future.result()
                clinics_raw.append(clinic)
                status = "✓ OPORTUNIDAD" if clinic["oportunidad"] else "  ok"
                print(f"      [{i}/{len(places)}] {status} - {clinic['nombre'][:50]} ({clinic['calidad_web']})")
            except Exception as e:
                place = futures[future]
                place_name = place.get("displayName", {})
                place_name = place_name.get("text", "?") if isinstance(place_name, dict) else "?"
                print(f"      [{i}/{len(places)}] ✗ Error procesando {place_name}: {e}")

    print(f"\n[3/4] Filtrando resultados...")
    clinics = []
    filtered_reviews = 0
    filtered_no_oportunidad = 0
    for c in clinics_raw:
        reviews = c["num_reviews"]
        review_count = int(reviews) if str(reviews).isdigit() else 0
        if review_count < args.min_reviews:
            filtered_reviews += 1
            continue
        if not c["oportunidad"]:
            filtered_no_oportunidad += 1
            continue
        clinics.append(c)

    print(f"      {len(clinics_raw)} clínicas analizadas")
    print(f"      {filtered_reviews} descartadas por menos de {args.min_reviews} reseñas")
    print(f"      {filtered_no_oportunidad} descartadas por tener buena web")
    print(f"      {len(clinics)} oportunidades válidas")

    existing_clinics = load_existing_clinics(args.output)
    if existing_clinics:
        print(f"\n      Archivo '{args.output}' ya existe con {len(existing_clinics)} clínicas. Fusionando...")
        all_clinics, added, skipped = merge_clinics(existing_clinics, clinics)
        print(f"      → {added} clínicas nuevas añadidas")
        print(f"      → {skipped} duplicadas ignoradas")
    else:
        all_clinics = clinics
        added = len(clinics)
        skipped = 0

    print(f"\n[4/4] Guardando resultados en {args.output}...")
    save_to_excel(all_clinics, args.output)

    sin_web = sum(1 for c in all_clinics if c["calidad_web"] == "Sin web")
    basica = sum(1 for c in all_clinics if "BÁSICA" in c["calidad_web"])

    print("\n" + "=" * 60)
    print("  RESUMEN")
    print("=" * 60)
    print(f"  Total en Excel:       {len(all_clinics)}")
    if existing_clinics:
        print(f"    (ya existían:       {len(existing_clinics)})")
        print(f"    (nuevas añadidas:   {added})")
        print(f"    (duplicadas:        {skipped})")
    print(f"  Sin web:              {sin_web}")
    print(f"  Web básica:           {basica}")
    print(f"\n  Archivo guardado en:  {args.output}")
    print("=" * 60)


if __name__ == "__main__":
    main()
