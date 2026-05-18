#!/usr/bin/env python3
"""
Buscador avanzado de emails para clínicas.

Lee el Excel generado por buscar_clinicas.py, identifica las clínicas
sin email, y usa múltiples estrategias para encontrar sus correos:

  1. Rastreo profundo de toda la web (hasta 40 páginas, emails ofuscados)
  2. Extracción de mailto:, JSON-LD, schema.org, meta tags
  3. Adivinación de patrones + verificación SMTP (todos los prefijos)
  4. Búsqueda en Google, DuckDuckGo y Bing
  5. Directorios: Doctoralia, Páginas Amarillas, TopDoctors, 11870, Cylex
  6. Redes sociales (Facebook, Instagram bio)
  7. Google Maps (ficha de la clínica)
  8. WHOIS del dominio

Uso:
    python buscar_emails.py
    python buscar_emails.py --input clinicas_barcelona.xlsx
    python buscar_emails.py --workers 8
"""

import argparse
import json
import re
import smtplib
import socket
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import quote_plus, urljoin, urlparse

import dns.resolver
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

HEADERS_BROWSER = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

OBFUSCATED_PATTERNS = [
    re.compile(r"([a-zA-Z0-9._%+\-]+)\s*[\[\(]\s*(?:arroba|at|@)\s*[\]\)]\s*([a-zA-Z0-9.\-]+)\s*[\[\(]\s*(?:punto|dot|\.)\s*[\]\)]\s*([a-zA-Z]{2,})", re.IGNORECASE),
    re.compile(r"([a-zA-Z0-9._%+\-]+)\s*(?:arroba|ARROBA)\s*([a-zA-Z0-9.\-]+)\s*(?:punto|PUNTO)\s*([a-zA-Z]{2,})", re.IGNORECASE),
    re.compile(r"([a-zA-Z0-9._%+\-]+)\s*\[at\]\s*([a-zA-Z0-9.\-]+)\s*\[dot\]\s*([a-zA-Z]{2,})", re.IGNORECASE),
]

JUNK_EXTENSIONS = (
    ".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp",
    ".css", ".js", ".woff", ".woff2", ".ttf", ".ico",
    ".pdf", ".zip", ".mp4", ".mp3",
)
JUNK_DOMAINS = (
    "sentry.io", "w3.org", "schema.org", "example.com",
    "wordpress.org", "gravatar.com", "googleapis.com",
    "googletagmanager.com", "facebook.com", "twitter.com",
    "instagram.com", "youtube.com", "linkedin.com",
    "cookiebot.com", "cookielaw.com", "onetrust.com",
    "cloudflare.com", "gstatic.com", "google.com",
    "wixpress.com", "squarespace.com", "wix.com",
    "hotjar.com", "hubspot.com", "mailchimp.com",
    "jsdelivr.net", "unpkg.com", "cdnjs.com",
)

CONTACT_KEYWORDS = [
    "contacto", "contact", "contacta", "contacte",
    "sobre-nosotros", "about", "quienes-somos", "quien-somos",
    "aviso-legal", "legal", "aviso_legal",
    "politica-de-privacidad", "privacidad", "privacy",
    "equipo", "team", "staff", "profesionales",
    "informacion", "info",
    "cita", "appointment", "pedir-cita", "reservar",
    "donde-estamos", "ubicacion", "como-llegar",
    "impressum", "imprint", "footer",
    "empresa", "nosotros", "clinica",
]

EMAIL_PREFIXES = [
    "info", "contacto", "contact", "clinica", "recepcion",
    "administracion", "admin", "hola", "consulta", "consultas",
    "citas", "atencion", "secretaria", "gerencia", "direccion",
    "recepcio", "consultes", "clínica",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def parse_args():
    parser = argparse.ArgumentParser(
        description="Busca emails de clínicas con múltiples estrategias."
    )
    parser.add_argument(
        "--input", "-i",
        default="clinicas_prospecto.xlsx",
        help="Archivo Excel (default: clinicas_prospecto.xlsx).",
    )
    parser.add_argument(
        "--workers", "-w",
        type=int,
        default=5,
        help="Hilos paralelos (default: 5).",
    )
    return parser.parse_args()


def clean_emails(raw_emails: list[str]) -> list[str]:
    """Filtra emails basura y devuelve lista única."""
    seen = set()
    cleaned = []
    for email in raw_emails:
        lower = email.lower().strip()
        if lower in seen:
            continue
        if lower.endswith(JUNK_EXTENSIONS):
            continue
        if any(domain in lower for domain in JUNK_DOMAINS):
            continue
        if len(lower) > 60 or len(lower) < 5:
            continue
        if lower.startswith(("noreply", "no-reply", "mailer-daemon", "postmaster")):
            continue
        if lower.count("@") != 1:
            continue
        domain_part = lower.split("@")[1]
        if "." not in domain_part:
            continue
        seen.add(lower)
        cleaned.append(email)
    return cleaned


def fetch_page(url: str, timeout: int = 10) -> str | None:
    """Descarga HTML de una URL."""
    try:
        resp = requests.get(
            url, headers=HEADERS_BROWSER,
            timeout=timeout, allow_redirects=True, verify=True,
        )
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if "text/html" not in content_type and "text/plain" not in content_type:
            return None
        return resp.text
    except requests.exceptions.SSLError:
        try:
            resp = requests.get(
                url, headers=HEADERS_BROWSER,
                timeout=timeout, allow_redirects=True, verify=False,
            )
            resp.raise_for_status()
            return resp.text
        except Exception:
            return None
    except Exception:
        return None


def extract_emails_from_html(html: str) -> list[str]:
    """Extrae emails del HTML: regex, mailto, ofuscados, JSON-LD, schema.org."""
    emails = []

    emails.extend(EMAIL_REGEX.findall(html))

    for pattern in OBFUSCATED_PATTERNS:
        for match in pattern.finditer(html):
            email = f"{match.group(1)}@{match.group(2)}.{match.group(3)}"
            emails.append(email)

    soup = BeautifulSoup(html, "html.parser")

    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        if href.startswith("mailto:"):
            email = href.replace("mailto:", "").split("?")[0].strip()
            if EMAIL_REGEX.match(email):
                emails.append(email)

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            emails.extend(_extract_emails_jsonld(data))
        except (json.JSONDecodeError, TypeError):
            pass

    for tag in soup.find_all(attrs={"itemprop": "email"}):
        content = tag.get("content", "") or tag.get_text(strip=True)
        if EMAIL_REGEX.match(content):
            emails.append(content)

    for meta in soup.find_all("meta"):
        content = meta.get("content", "")
        if "@" in content:
            emails.extend(EMAIL_REGEX.findall(content))

    return emails


def _extract_emails_jsonld(data) -> list[str]:
    """Recorre recursivamente JSON-LD buscando emails."""
    emails = []
    if isinstance(data, dict):
        for key, value in data.items():
            if key.lower() in ("email", "contactpoint", "contactpoints"):
                if isinstance(value, str) and EMAIL_REGEX.match(value):
                    emails.append(value)
                elif isinstance(value, (list, dict)):
                    emails.extend(_extract_emails_jsonld(value))
            else:
                emails.extend(_extract_emails_jsonld(value))
    elif isinstance(data, list):
        for item in data:
            emails.extend(_extract_emails_jsonld(item))
    return emails


# ---------------------------------------------------------------------------
# Search engines
# ---------------------------------------------------------------------------


def _search_google(query: str, num: int = 5) -> list[str]:
    """Google search, devuelve URLs."""
    try:
        resp = requests.get(
            "https://www.google.com/search",
            params={"q": query, "num": num, "hl": "es"},
            headers={**HEADERS_BROWSER, "Referer": "https://www.google.com/"},
            timeout=10,
        )
        if resp.status_code != 200:
            return []
        urls = []
        for a in BeautifulSoup(resp.text, "html.parser").find_all("a", href=True):
            href = a["href"]
            if href.startswith("/url?q="):
                real = href.split("/url?q=")[1].split("&")[0]
                if real.startswith("http") and "google." not in real:
                    urls.append(real)
        return urls[:num]
    except Exception:
        return []


def _search_duckduckgo(query: str, num: int = 5) -> list[str]:
    """DuckDuckGo HTML search, devuelve URLs."""
    try:
        resp = requests.get(
            "https://html.duckduckgo.com/html/",
            params={"q": query},
            headers=HEADERS_BROWSER,
            timeout=10,
        )
        if resp.status_code != 200:
            return []
        urls = []
        soup = BeautifulSoup(resp.text, "html.parser")
        for a in soup.find_all("a", class_="result__a", href=True):
            href = a["href"]
            if href.startswith("http"):
                urls.append(href)
        return urls[:num]
    except Exception:
        return []


def _search_bing(query: str, num: int = 5) -> list[str]:
    """Bing search, devuelve URLs."""
    try:
        resp = requests.get(
            "https://www.bing.com/search",
            params={"q": query, "count": num},
            headers=HEADERS_BROWSER,
            timeout=10,
        )
        if resp.status_code != 200:
            return []
        urls = []
        soup = BeautifulSoup(resp.text, "html.parser")
        for li in soup.find_all("li", class_="b_algo"):
            a = li.find("a", href=True)
            if a and a["href"].startswith("http"):
                urls.append(a["href"])
        return urls[:num]
    except Exception:
        return []


def multi_search(query: str, num: int = 5) -> list[str]:
    """Busca en Google, DuckDuckGo y Bing. Combina resultados."""
    seen = set()
    results = []
    for search_fn in [_search_google, _search_duckduckgo, _search_bing]:
        for url in search_fn(query, num):
            if url not in seen:
                seen.add(url)
                results.append(url)
        if len(results) >= num:
            break
    return results[:num * 2]


def search_and_extract_emails(query: str, max_pages: int = 5) -> list[str]:
    """Busca en buscadores y extrae emails de las páginas resultantes."""
    emails = []
    urls = multi_search(query, num=max_pages)
    for url in urls[:max_pages]:
        html = fetch_page(url, timeout=8)
        if not html:
            continue
        found = extract_emails_from_html(html)
        emails.extend(found)
        if clean_emails(emails):
            break
    return clean_emails(emails)


# ---------------------------------------------------------------------------
# Strategy 1: Deep website crawl
# ---------------------------------------------------------------------------


def crawl_website(base_url: str, max_pages: int = 40) -> tuple[list[str], list[str]]:
    """Rastrea todo el sitio web buscando emails. Prioriza contacto."""
    if not base_url:
        return [], []

    parsed_base = urlparse(base_url)
    base_domain = parsed_base.netloc
    visited = set()
    all_emails = []
    pages_visited = []
    priority_queue = []
    normal_queue = []

    main_html = fetch_page(base_url)
    if not main_html:
        return [], [base_url]

    visited.add(base_url)
    pages_visited.append(base_url)
    all_emails.extend(extract_emails_from_html(main_html))

    soup = BeautifulSoup(main_html, "html.parser")
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"].strip()
        if not href or href.startswith(("#", "javascript:", "tel:")):
            continue

        if href.startswith("mailto:"):
            email = href.replace("mailto:", "").split("?")[0].strip()
            if EMAIL_REGEX.match(email):
                all_emails.append(email)
            continue

        full_url = urljoin(base_url, href)
        parsed = urlparse(full_url)
        if parsed.netloc and parsed.netloc != base_domain:
            continue
        if any(full_url.lower().endswith(ext) for ext in JUNK_EXTENSIONS):
            continue

        clean_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
        if clean_url in visited:
            continue
        visited.add(clean_url)

        path_lower = parsed.path.lower()
        link_text = a_tag.get_text(strip=True).lower()
        is_priority = any(kw in path_lower or kw in link_text for kw in CONTACT_KEYWORDS)

        if is_priority:
            priority_queue.append(clean_url)
        else:
            normal_queue.append(clean_url)

    for url in priority_queue + normal_queue:
        if len(pages_visited) >= max_pages:
            break
        html = fetch_page(url, timeout=8)
        if not html:
            continue
        pages_visited.append(url)
        page_emails = extract_emails_from_html(html)
        all_emails.extend(page_emails)

    return clean_emails(all_emails), pages_visited


# ---------------------------------------------------------------------------
# Strategy 2: Email pattern guessing + SMTP verify
# ---------------------------------------------------------------------------


def get_domain_from_web(web_url: str) -> str | None:
    """Extrae el dominio raíz de una URL."""
    if not web_url:
        return None
    parsed = urlparse(web_url)
    domain = parsed.netloc
    if domain.startswith("www."):
        domain = domain[4:]
    return domain if domain else None


def check_mx_exists(domain: str) -> bool:
    """Verifica si un dominio tiene registros MX."""
    try:
        dns.resolver.resolve(domain, "MX")
        return True
    except Exception:
        return False


def verify_email_smtp(email: str, mx_host: str) -> bool | None:
    """
    Verifica si un email existe via SMTP.
    Retorna True (existe), False (no existe), None (no se pudo verificar).
    """
    try:
        smtp = smtplib.SMTP(timeout=5)
        smtp.connect(mx_host, 25)
        smtp.helo("check.local")
        smtp.mail("test@check.local")
        code, _ = smtp.rcpt(email)
        smtp.quit()
        if code == 250:
            return True
        elif code == 550 or code == 551 or code == 553:
            return False
        return None
    except Exception:
        return None


def guess_and_verify_emails(web_url: str) -> list[str]:
    """Genera emails probables y verifica cuáles existen via SMTP."""
    domain = get_domain_from_web(web_url)
    if not domain:
        return []

    if not check_mx_exists(domain):
        return []

    try:
        records = dns.resolver.resolve(domain, "MX")
        mx_host = str(records[0].exchange).rstrip(".")
    except Exception:
        return []

    verified = []
    catch_all = None

    for prefix in EMAIL_PREFIXES:
        candidate = f"{prefix}@{domain}"
        result = verify_email_smtp(candidate, mx_host)
        if result is True:
            verified.append(candidate)
            if len(verified) >= 3:
                catch_all = True
                break
        elif result is False:
            continue

    if catch_all:
        return [f"info@{domain}"]

    if not verified:
        random_test = f"xq9z8w7test@{domain}"
        result = verify_email_smtp(random_test, mx_host)
        if result is True:
            return [f"info@{domain}"]

    return verified[:2]


# ---------------------------------------------------------------------------
# Strategy 3: Search engines
# ---------------------------------------------------------------------------


def search_email_engines(name: str, city: str) -> list[str]:
    """Busca el email con múltiples buscadores y consultas."""
    queries = [
        f'"{name}" correo OR email OR "@"',
    ]
    if city:
        queries.append(f'"{name}" "{city}" email')

    for query in queries:
        emails = search_and_extract_emails(query, max_pages=4)
        if emails:
            return emails
    return []


# ---------------------------------------------------------------------------
# Strategy 4: Spanish directories (direct scrape)
# ---------------------------------------------------------------------------


DIRECTORY_SEARCHES = [
    'site:doctoralia.es "{name}"',
    'site:topdoctors.es "{name}"',
    'site:paginasamarillas.es "{name}"',
    'site:11870.com "{name}"',
    'site:cylex.es "{name}"',
    'site:infoisinfo.es "{name}"',
    'site:clinicasesteticas.com "{name}"',
]


def search_directories(name: str, city: str) -> list[str]:
    """Busca en directorios españoles vía buscadores."""
    all_emails = []

    for template in DIRECTORY_SEARCHES:
        query = template.format(name=name)
        urls = multi_search(query, num=3)
        for url in urls[:2]:
            html = fetch_page(url, timeout=8)
            if not html:
                continue
            found = extract_emails_from_html(html)
            cleaned = clean_emails(found)
            if cleaned:
                return cleaned

    return clean_emails(all_emails)


# ---------------------------------------------------------------------------
# Strategy 5: Social media
# ---------------------------------------------------------------------------


def find_social_email(html: str) -> list[str]:
    """Busca emails en páginas de redes sociales enlazadas."""
    soup = BeautifulSoup(html, "html.parser")
    social_urls = []

    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        if "facebook.com" in href and "/posts/" not in href and "/photos/" not in href:
            social_urls.append(href)
        elif "instagram.com" in href:
            social_urls.append(href)

    emails = []
    for url in social_urls[:3]:
        page_html = fetch_page(url, timeout=8)
        if not page_html:
            continue
        found = EMAIL_REGEX.findall(page_html)
        emails.extend(found)
        if clean_emails(emails):
            break

    return clean_emails(emails)


# ---------------------------------------------------------------------------
# Strategy 6: Google Maps page
# ---------------------------------------------------------------------------


def search_google_maps_email(name: str, city: str) -> list[str]:
    """Busca el email en la ficha de Google Maps."""
    query = f'"{name}" {city} site:google.com/maps'
    urls = _search_google(query, num=3)
    emails = []
    for url in urls[:2]:
        html = fetch_page(url, timeout=8)
        if not html:
            continue
        found = EMAIL_REGEX.findall(html)
        emails.extend(found)
    return clean_emails(emails)


# ---------------------------------------------------------------------------
# Strategy 7: WHOIS
# ---------------------------------------------------------------------------


def whois_email(web_url: str) -> list[str]:
    """Intenta obtener email del WHOIS del dominio."""
    domain = get_domain_from_web(web_url)
    if not domain:
        return []

    try:
        resp = requests.get(
            f"https://whois.domaintools.com/{domain}",
            headers=HEADERS_BROWSER,
            timeout=8,
        )
        if resp.status_code != 200:
            return []
        found = EMAIL_REGEX.findall(resp.text)
        return clean_emails(found)
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


STRATEGY_ORDER = [
    ("Rastreo web profundo", "crawl"),
    ("Redes sociales", "social"),
    ("Patrón email + SMTP", "smtp"),
    ("Buscadores (Google/DDG/Bing)", "search"),
    ("Directorios españoles", "directories"),
    ("Google Maps", "gmaps"),
    ("WHOIS", "whois"),
]


def find_email_for_clinic(name: str, web: str, address: str) -> dict:
    """Ejecuta todas las estrategias en cascada para encontrar el email."""
    result = {
        "name": name,
        "emails": [],
        "strategy_used": None,
        "pages_crawled": 0,
        "strategies_tried": [],
    }

    city = ""
    if address:
        parts = [p.strip() for p in address.split(",")]
        for part in parts:
            cleaned = re.sub(r"\d{5}", "", part).strip()
            if cleaned and len(cleaned) > 2 and not cleaned.isdigit():
                city = cleaned
                break

    cached_main_html = None

    for strategy_name, strategy_id in STRATEGY_ORDER:
        result["strategies_tried"].append(strategy_id)
        emails = []

        try:
            if strategy_id == "crawl" and web:
                emails, pages = crawl_website(web, max_pages=40)
                result["pages_crawled"] = len(pages)
                if not emails and pages:
                    cached_main_html = fetch_page(web)

            elif strategy_id == "social" and web:
                html = cached_main_html or fetch_page(web)
                if html:
                    cached_main_html = html
                    emails = find_social_email(html)

            elif strategy_id == "smtp" and web:
                emails = guess_and_verify_emails(web)

            elif strategy_id == "search":
                emails = search_email_engines(name, city)

            elif strategy_id == "directories":
                emails = search_directories(name, city)

            elif strategy_id == "gmaps":
                emails = search_google_maps_email(name, city)

            elif strategy_id == "whois" and web:
                emails = whois_email(web)

        except Exception:
            continue

        if emails:
            result["emails"] = emails
            result["strategy_used"] = strategy_name
            return result

    result["strategy_used"] = "No encontrado"
    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    args = parse_args()

    print("=" * 60)
    print("  BUSCADOR AVANZADO DE EMAILS")
    print("  Sistema multi-estrategia (7 métodos)")
    print("=" * 60)

    try:
        wb = load_workbook(args.input)
    except FileNotFoundError:
        print(f"\n  [ERROR] No se encontró '{args.input}'")
        print("  Ejecuta primero buscar_clinicas.py para generar el Excel.")
        sys.exit(1)
    except Exception as e:
        print(f"\n  [ERROR] No se pudo abrir '{args.input}': {e}")
        sys.exit(1)

    ws = wb["Clínicas - Prospectos"]

    clinics_to_search = []
    total_clinics = 0
    already_have_email = 0

    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        if not name:
            continue
        total_clinics += 1
        email = ws.cell(row=row_idx, column=3).value
        address = ws.cell(row=row_idx, column=4).value or ""

        web_cell = ws.cell(row=row_idx, column=5)
        web = ""
        if web_cell.hyperlink:
            web = web_cell.hyperlink.target
        elif web_cell.value and web_cell.value != "Ver web":
            web = web_cell.value

        if email and str(email).strip():
            already_have_email += 1
        else:
            clinics_to_search.append({
                "row": row_idx,
                "name": str(name),
                "web": web,
                "address": str(address),
            })

    print(f"\n  Total clínicas en Excel:    {total_clinics}")
    print(f"  Ya tienen email:            {already_have_email}")
    print(f"  Sin email (a buscar):       {len(clinics_to_search)}")
    sin_web = sum(1 for c in clinics_to_search if not c["web"])
    con_web = len(clinics_to_search) - sin_web
    print(f"    - Con web (7 estrategias):         {con_web}")
    print(f"    - Sin web (buscadores/directorios): {sin_web}")

    if not clinics_to_search:
        print("\n  Todas las clínicas ya tienen email.")
        wb.close()
        sys.exit(0)

    print(f"\n  Estrategias (en orden):")
    print(f"    1. Rastreo web profundo (hasta 40 págs, emails ofuscados)")
    print(f"    2. Redes sociales (Facebook, Instagram)")
    print(f"    3. Adivinar patrón + verificar SMTP (15 prefijos)")
    print(f"    4. Buscadores (Google + DuckDuckGo + Bing)")
    print(f"    5. Directorios (Doctoralia, PáginasAmarillas, TopDoctors...)")
    print(f"    6. Ficha de Google Maps")
    print(f"    7. WHOIS del dominio")
    print(f"\n  Buscando ({args.workers} hilos)...\n")

    found_count = 0
    not_found_count = 0
    strategy_stats = {}
    results = {}

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(
                find_email_for_clinic,
                c["name"], c["web"], c["address"]
            ): c
            for c in clinics_to_search
        }
        for i, future in enumerate(as_completed(futures), 1):
            clinic = futures[future]
            try:
                res = future.result()
                short_name = res["name"][:40]
                if res["emails"]:
                    found_count += 1
                    results[clinic["row"]] = res["emails"]
                    strategy = res["strategy_used"]
                    strategy_stats[strategy] = strategy_stats.get(strategy, 0) + 1
                    emails_str = ", ".join(res["emails"][:3])
                    print(f"  [{i}/{len(clinics_to_search)}] ENCONTRADO  {short_name}")
                    print(f"       → {emails_str}")
                    print(f"       Via: {strategy}")
                else:
                    not_found_count += 1
                    tried = len(res["strategies_tried"])
                    print(f"  [{i}/{len(clinics_to_search)}] sin resultado  {short_name}  ({tried} estrategias, {res['pages_crawled']} págs)")
            except Exception as e:
                not_found_count += 1
                print(f"  [{i}/{len(clinics_to_search)}] ERROR  {clinic['name'][:40]}: {e}")

    if found_count > 0:
        print(f"\n  Actualizando Excel con {found_count} emails...")
        for row_idx, emails in results.items():
            ws.cell(row=row_idx, column=3, value=", ".join(emails))
        wb.save(args.input)
        print(f"  Guardado: {args.input}")
    else:
        print(f"\n  No se encontraron emails nuevos.")

    print("\n" + "=" * 60)
    print("  RESUMEN")
    print("=" * 60)
    print(f"  Clínicas buscadas:      {len(clinics_to_search)}")
    print(f"  Emails encontrados:     {found_count}")
    print(f"  Sin resultado:          {not_found_count}")
    if strategy_stats:
        print(f"\n  Emails por estrategia:")
        for strategy, count in sorted(strategy_stats.items(), key=lambda x: -x[1]):
            print(f"    {strategy}: {count}")
    print("=" * 60)

    wb.close()


if __name__ == "__main__":
    main()
